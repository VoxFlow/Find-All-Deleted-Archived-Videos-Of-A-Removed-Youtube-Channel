import subprocess
import json
import os
import time
import signal
import sys
import shutil
from openpyxl import Workbook, load_workbook

INPUT_FILE = "Downloads.txt"
OUTPUT_FILE = "Titles+Dates+Uploader.xlsx"

SLEEP_SECONDS = 5
RETRY_DELAY = 10
MAX_RETRIES = 3

wb = None
ws = None


# -------------------- Auto-merge tmp file on startup --------------------

def auto_merge_tmp():
    tmp_file = OUTPUT_FILE + ".tmp"
    if os.path.exists(tmp_file):
        if (not os.path.exists(OUTPUT_FILE) or
            os.path.getmtime(tmp_file) > os.path.getmtime(OUTPUT_FILE)):
            print(f"Found newer temp file '{tmp_file}'. Restoring as main output.")
            try:
                shutil.move(tmp_file, OUTPUT_FILE)
                print("Auto-merge complete.")
            except Exception as e:
                print(f"Failed to auto-merge temp file: {e}")
        else:
            # Temp file exists but is older; delete it
            try:
                os.remove(tmp_file)
            except Exception:
                pass


# -------------------- yt-dlp --------------------

def run_yt_dlp(url, retries=MAX_RETRIES):
    cmd = [
        "yt-dlp",
        "--skip-download",
        "--dump-json",
        "--no-playlist",
        "--ignore-errors",
        "--no-warnings",
        url
    ]

    for attempt in range(1, retries + 1):
        try:
            result = subprocess.run(
                cmd,
                capture_output=True,
                text=True,
                check=True
            )
            return json.loads(result.stdout)
        except Exception as e:
            print(f"  yt-dlp failed (attempt {attempt}). Retrying in {RETRY_DELAY}s...")
            time.sleep(RETRY_DELAY)

    return None


# -------------------- XLSX helpers --------------------

def load_or_create_workbook():
    if os.path.exists(OUTPUT_FILE):
        wb = load_workbook(OUTPUT_FILE)
        ws = wb.active
    else:
        wb = Workbook()
        ws = wb.active
        ws.title = "Titles and Dates"
        ws.append(["video_id", "title", "upload_date", "uploader", "source_url"])
        wb.save(OUTPUT_FILE)
    return wb, ws


def get_processed_urls(ws):
    processed = set()
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row and row[4]:
            processed.add(row[4])
    return processed


def atomic_save(workbook):
    tmp = OUTPUT_FILE + ".tmp"
    workbook.save(tmp)

    try:
        os.replace(tmp, OUTPUT_FILE)
    except PermissionError:
        print("Excel has the file open — saved progress to temp file instead.")


# -------------------- SIGINT handler --------------------

def handle_sigint(signum, frame):
    print("\nCtrl+C detected. Saving progress...")
    try:
        atomic_save(wb)
        print("Progress saved. Exiting cleanly.")
    except Exception as e:
        print(f"Failed to save workbook: {e}")
    sys.exit(0)


# -------------------- main --------------------

def main():
    global wb, ws

    auto_merge_tmp()

    if not os.path.exists(INPUT_FILE):
        print(f"Input file '{INPUT_FILE}' not found.")
        return

    with open(INPUT_FILE, encoding="utf-8") as f:
        urls = [line.strip() for line in f if line.strip()]

    wb, ws = load_or_create_workbook()
    signal.signal(signal.SIGINT, handle_sigint)

    processed_urls = get_processed_urls(ws)

    print(f"Already processed: {len(processed_urls)} URLs\n")

    for idx, url in enumerate(urls, 1):
        if url in processed_urls:
            print(f"[SKIP] {url}")
            continue

        print(f"[{idx}/{len(urls)}] Processing {url}")

        data = run_yt_dlp(url)

        if not data:
            ws.append([
                "unknown",
                "unavailable",
                "unavailable",
                "unavailable",
                url
            ])
            atomic_save(wb)
            time.sleep(SLEEP_SECONDS)
            continue

        video_id = data.get("id", "unknown")
        title = data.get("title", "unavailable")
        uploader = data.get("uploader", "unavailable")

        raw_date = data.get("upload_date", "")
        if len(raw_date) == 8:
            upload_date = f"{raw_date[:4]}-{raw_date[4:6]}-{raw_date[6:]}"
        else:
            upload_date = "unavailable"

        print(f"  Title: {title}")
        print(f"  Date: {upload_date}")
        print(f"  Uploader: {uploader}")

        ws.append([
            video_id,
            title,
            upload_date,
            uploader,
            url
        ])

        atomic_save(wb)
        time.sleep(SLEEP_SECONDS)

    print("\nAll URLs processed.")


if __name__ == "__main__":
    main()
